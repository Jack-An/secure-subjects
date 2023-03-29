import json

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


class SubjectParser:

    @classmethod
    def single_subject_paser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 5):
            subject_rows = rows[i:i + 5]
            first_line = subject_rows[0]
            tag = first_line[1].value
            title = first_line[3].value
            answer = first_line[4].value
            options = []
            for option in subject_rows[1:]:
                option_val = option[3].value
                splits = option_val.split('.')
                key = splits[0]
                text = "".join(splits[1:])
                options.append({"key": key, "option": text})
            subject = {
                "tag": tag,
                "title": title,
                "answer": answer,
                "options": options
            }
            subjects.append(subject)
        return subjects

    @classmethod
    def multiple_subject_parser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 6):
            subject_rows = rows[i:i + 6]
            first_line = subject_rows[0]
            tag = first_line[1].value
            title = first_line[3].value
            answer = first_line[4].value
            answer = [val for val in answer]
            options = []
            for option in subject_rows[1:]:
                option_val = option[3].value
                splits = option_val.split('.')
                key = splits[0]
                text = "".join(splits[1:])
                options.append({"key": key, "option": text})
            subject = {
                "tag": tag,
                "title": title,
                "answer": answer,
                "options": options
            }
            subjects.append(subject)
        return subjects

    @classmethod
    def judge_subject_parser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 3):
            subject_rows = rows[i:i + 3]
            first_line = subject_rows[0]
            tag = first_line[1].value
            title = first_line[3].value
            answer = first_line[4].value
            options = []
            for option in subject_rows[1:]:
                option_val = option[3].value
                splits = option_val.split('.')
                key = splits[0]
                text = "".join(splits[1:])
                options.append({"key": key, "option": text})
            subject = {
                "tag": tag,
                "title": title,
                "answer": answer,
                "options": options
            }
            subjects.append(subject)
        return subjects

    @classmethod
    def _split_sample_subjects(cls, rows):
        subjects = []
        start_idx = 0
        for i in range(1, len(rows)):
            if rows[i][0].value:
                end_idx = i
                subjects.append(rows[start_idx:end_idx])
                start_idx = i + 1
        subjects.append(rows[start_idx+1:])
        return subjects


    @classmethod
    def sample_subject_parser(cls, rows):
        subjects = []
        for i in range(0, len(rows), 3):
            subject_rows = rows[i:i + 3]
            first_line = subject_rows[0]
            tag = first_line[1].value
            title = first_line[3].value
            answer = first_line[4].value
            options = []
            for option in subject_rows[1:]:
                option_val = option[3].value
                splits = option_val.split('.')
                key = splits[0]
                text = "".join(splits[1:])
                options.append({"key": key, "option": text})
            subject = {
                "tag": tag,
                "title": title,
                "answer": answer,
                "options": options
            }
            subjects.append(subject)
        return subjects


    @classmethod
    def clear_rows(cls, rows):
        valid_rows = []
        for k in rows:
            if k[3].value and k[3].value.strip():
                valid_rows.append(k)
        return valid_rows

    @classmethod
    def read_excel(cls):
        wb: Workbook = load_workbook(filename="../lib/建设工程安全生产管理.xlsx")
        sheet_key_map = {
            '单选': "single",
            "多选": "multiple",
            "判断": "judge",
            "案例": "sample"
        }
        result = {}
        for sheet_name in wb.sheetnames:
            subject_key = sheet_key_map[sheet_name]
            table = wb[sheet_name]
            rows = [row for row in table.rows][1:]
            valid_rows = cls.clear_rows(rows)
            if subject_key == "single":
                subjects = cls.single_subject_paser(valid_rows)
            elif subject_key == "multiple":
                subjects = cls.multiple_subject_parser(valid_rows)
            elif subject_key == "judge":
                subjects = cls.judge_subject_parser(valid_rows)
            else:
                subjects = cls._split_sample_subjects(valid_rows)
                print("sample total", len(subjects))
                print(subjects[-1])
                subjects = []

            with open(f'{subject_key}.json', 'w') as f:
                json.dump(subjects, f)
            # print(len(subjects))
            # print(subjects[:3])







if __name__ == '__main__':
    SubjectParser.read_excel()



